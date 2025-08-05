import openpyxl as xl
from openpyxl.chart import BarChart, Reference

class Work:
	# @staticmethod
	def workbook(name):
		wb = xl.load_workbook(name)
		sheet = wb['Sheet1']

		for row in range(2, sheet.max_row + 1):
			cell = sheet.cell(row, 3)
			cell_p = cell.value * 1.9
			cell_c = sheet.cell(row, 3)
			cell_c.value = cell_p

			values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col = 4, max_col = 4)

			chart = BarChart()
			chart.add_data(values)

			sheet.add_chart(chart, 'f2')
			wb.save('transactions.xlsx')